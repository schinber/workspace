# 在app/routes.py顶部导入必要模块
import logging

from datetime import datetime
from functools import wraps

from flask import Blueprint, render_template, current_app
from flask import redirect, url_for, flash, request
from flask import session
from flask_login import login_user, logout_user, login_required, current_user
from openpyxl import load_workbook

from app import db
from app.models import User, Student
from app.utils import role_required, is_safe_url

main = Blueprint('main', __name__)


logger = logging.getLogger(__name__)

# 新增：会话超时检查装饰器
def check_session_timeout(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        # 仅对已登录用户生效
        if current_user.is_authenticated:
            # 检查会话是否过期
            if 'login_time' in session:
                login_time = datetime.strptime(session['login_time'], "%Y-%m-%d %H:%M:%S")
                # 计算当前时间与登录时间的差值
                time_diff = datetime.now() - login_time
                # 如果超过配置的超时时间，强制登出
                if time_diff > current_app.config['PERMANENT_SESSION_LIFETIME']:
                    logout_user()
                    session.clear()
                    flash('由于长时间未操作，您已被自动登出，请重新登录')
                    return redirect(url_for('main.login', next=request.url))
            else:
                # 如果没有登录时间记录，强制登出
                logout_user()
                flash('会话异常，请重新登录')
                return redirect(url_for('main.login'))
        return f(*args, **kwargs)

    return decorated_function


@main.route('/keep_alive')
@login_required
def keep_alive():
    # 更新会话中的登录时间
    session['login_time'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    return 'OK'


@main.route('/')
def index():
    if current_user.is_authenticated:
        if current_user.is_admin():
            return redirect(url_for('main.admin_dashboard'))
        elif current_user.is_teacher():
            return redirect(url_for('main.teacher_dashboard'))
        elif current_user.is_student():
            return redirect(url_for('main.student_profile'))
    return redirect(url_for('main.login'))


@main.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('main.index'))

    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        user = User.query.filter_by(username=username).first()

        if user and user.check_password(password):
            login_user(user)
            session.permanent = True
            session['login_time'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            # 获取原访问地址（next参数），若合法则跳转，否则跳转到首页
            next_page = request.args.get('next')
            if next_page and is_safe_url(next_page):  # 验证next参数安全性
                return redirect(next_page)
            return redirect(url_for('main.index'))
        else:
            flash('Invalid username or password')

    return render_template('login.html')


@main.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('main.login'))


# 管理员路由
@main.route('/admin/dashboard')
@login_required
@check_session_timeout
@role_required('admin')  # 仅允许管理员访问
def admin_dashboard():
    # ... 直接写业务逻辑，无需再判断权限 ...
    total_students = Student.query.count()
    total_teachers = User.query.filter_by(role='teacher').count()
    total_users = User.query.count()
    return render_template('admin/dashboard.html',
                           total_students=total_students,
                           total_teachers=total_teachers,
                           total_users=total_users)


@main.route('/admin/manage_users', methods=['GET', 'POST'])
@login_required
def manage_users():
    if not current_user.is_admin():
        flash('Access denied')
        return redirect(url_for('main.index'))

    if request.method == 'POST':
        action = request.form.get('action')
        user_id = request.form.get('user_id')

        if action == 'delete' and user_id:
            user = User.query.get(user_id)
            if user and user.username != 'admin':  # 不能删除admin用户
                db.session.delete(user)
                db.session.commit()
                flash('User deleted successfully')
            elif user and user.username == 'admin':
                flash('Cannot delete admin user')

        elif action == 'add':
            username = request.form.get('new_username')
            password = request.form.get('new_password')
            role = request.form.get('new_role')

            if User.query.filter_by(username=username).first():
                flash('Username already exists')
            else:
                user = User(username=username, role=role)
                user.set_password(password)
                db.session.add(user)
                db.session.commit()
                flash('User added successfully')

    users = User.query.all()
    return render_template('admin/manage_users.html', users=users)


# 管理员管理学生路由（添加权限控制）
# @main.route('/admin/manage_students', methods=['GET', 'POST'])
# @login_required  # 要求登录
# @check_session_timeout  # 会话超时检查（如果已实现）
# @role_required('admin')
# def manage_students():
#     # if not current_user.is_admin():
#     #     flash('Access denied')
#     #     return redirect(url_for('main.index'))
#











#
#         # 新增：批量导入逻辑
#         if action == 'batch_import':
#             # 检查是否有文件上传
#             if 'excel_file' not in request.files:
#                 flash('请选择Excel文件')
#                 return redirect(request.url)
#
#             file = request.files['excel_file']
#             if file.filename == '':
#                 flash('未选择文件')
#                 return redirect(request.url)
#
#             # 检查文件格式
#             if not file.filename.endswith('.xlsx'):
#                 flash('请上传.xlsx格式的Excel文件')
#                 return redirect(request.url)
#
#             try:
#                 # 加载Excel文件
#                 workbook = load_workbook(file.stream)
#                 sheet = workbook.active  # 获取第一个工作表
#
#                 # 检查表头（第一行应为标题行）
#                 headers = [cell.value for cell in sheet[1]]
#                 required_headers = ['姓名', '学号', '用户名', '密码']  # 必需的列
#                 for header in required_headers:
#                     if header not in headers:
#                         flash(f'Excel文件缺少必需的列：{header}')
#                         return redirect(request.url)
#
#                 # 解析数据行（从第二行开始）
#                 success_count = 0
#                 fail_count = 0
#                 fail_reasons = []
#
#                 for row in sheet.iter_rows(min_row=2, values_only=True):
#                     # 将行数据转换为字典（表头->值）
#                     row_data = dict(zip(headers, row))
#
#                     # 提取必要字段
#                     name = row_data.get('姓名')
#                     student_number = row_data.get('学号')
#                     username = row_data.get('用户名')
#                     password = row_data.get('密码')
#
#                     # 验证必要字段
#                     if not all([name, student_number, username, password]):
#                         fail_count += 1
#                         fail_reasons.append(f'行{row[0].row}：缺少必要信息')
#                         continue
#
#                     # 检查学号和用户名是否已存在
#                     if Student.query.filter_by(student_number=student_number).first():
#                         fail_count += 1
#                         fail_reasons.append(f'行{row[0]}：学号{student_number}已存在')
#                         continue
#
#                     if User.query.filter_by(username=username).first():
#                         fail_count += 1
#                         fail_reasons.append(f'行{row[1]}：用户名{username}已存在')
#                         continue
#                     logging.info("开始创建学生记录")
#                     # 创建学生记录
#                     student = Student(
#                         name=name,
#                         student_number=student_number,
#                         gender=row_data.get('性别'),  # 可选字段
#                         age=row_data.get('年龄'),  # 可选字段
#                         major=row_data.get('专业'),  # 可选字段
#                         email=row_data.get('邮箱')  # 可选字段
#                     )
#                     db.session.add(student)
#                     db.session.commit()  # 提交以获取student.id
#                     logging.info("创建关联的用户")
#                     # 创建关联的用户
#                     user = User(
#                         username=username,
#                         role='student',
#                         student_id=student.id
#                     )
#                     user.set_password(str(password))
#                     db.session.add(user)
#                     db.session.commit()
#
#                     success_count += 1
#
#                 # 显示导入结果
#                 flash(f'批量导入完成！成功：{success_count} 条，失败：{fail_count} 条')
#                 if fail_reasons:
#                     for reason in fail_reasons:
#                         flash(f'失败原因：{reason}')
#
#             except Exception as e:
#                 db.session.rollback()
#                 flash(f'导入失败：{str(e)}')
#             return redirect(url_for('main.manage_students'))
#
#         if action == 'delete' and student_id:
#             student = Student.query.get(student_id)
#             if student:
#                 # 删除关联的用户
#                 user = User.query.filter_by(student_id=student_id).first()
#                 if user:
#                     db.session.delete(user)
#                 db.session.delete(student)
#                 db.session.commit()
#                 flash('Student deleted successfully')
#
#         elif action == 'add':
#             name = request.form.get('name')
#             student_number = request.form.get('student_number')
#             gender = request.form.get('gender')
#             age = request.form.get('age')
#             major = request.form.get('major')
#             email = request.form.get('email')
#             username = request.form.get('username')
#             password = request.form.get('password')
#
#             if Student.query.filter_by(student_number=student_number).first():
#                 flash('Student number already exists')
#             elif User.query.filter_by(username=username).first():
#                 flash('Username already exists')
#             else:
#                 student = Student(
#                     name=name,
#                     student_number=student_number,
#                     gender=gender,
#                     age=int(age) if age else None,
#                     major=major,
#                     email=email
#                 )
#                 db.session.add(student)
#                 db.session.commit()
#
#                 # 创建学生用户
#                 user = User(
#                     username=username,
#                     role='student',
#                     student_id=student.id
#                 )
#                 user.set_password(password)
#                 db.session.add(user)
#                 db.session.commit()
#
#                 flash('Student added successfully')
#
#         elif action == 'edit' and student_id:
#             student = Student.query.get(student_id)
#             if student:
#                 student.name = request.form.get('name')
#                 student.student_number = request.form.get('student_number')
#                 student.gender = request.form.get('gender')
#                 student.age = int(request.form.get('age')) if request.form.get('age') else None
#                 student.major = request.form.get('major')
#                 student.email = request.form.get('email')
#                 db.session.commit()
#                 flash('Student updated successfully')
#
#     students = Student.query.all()
#     return render_template('admin/manage_students.html', students=students)


# 管理员管理学生路由（添加权限控制）
@main.route('/admin/manage_students', methods=['GET', 'POST'])
@login_required  # 要求登录
@check_session_timeout  # 会话超时检查（如果已实现）
@role_required('admin')
def manage_students():
    # if not current_user.is_admin():
    #     flash('Access denied')
    #     return redirect(url_for('main.index'))

    if request.method == 'POST':
        action = request.form.get('action')
        student_id = request.form.get('student_id')
        # 新增：批量导入逻辑
        if action == 'batch_import':
            # 检查是否有文件上传
            if 'excel_file' not in request.files:
                flash('请选择Excel文件')
                return redirect(request.url)

            file = request.files['excel_file']
            if file.filename == '':
                flash('未选择文件')
                return redirect(request.url)

            # 检查文件格式
            if not file.filename.endswith('.xlsx'):
                flash('请上传.xlsx格式的Excel文件')
                return redirect(request.url)

            try:
                # 加载Excel文件
                workbook = load_workbook(file.stream)
                sheet = workbook.active  # 获取第一个工作表

                # 检查表头（第一行应为标题行）
                headers = [cell.value for cell in sheet[1]]
                required_headers = ['姓名', '学号', '用户名', '密码']  # 必需的列
                for header in required_headers:
                    if header not in headers:
                        flash(f'Excel文件缺少必需的列：{header}')
                        return redirect(request.url)

                # 解析数据行（从第二行开始）
                success_count = 0
                fail_count = 0
                fail_reasons = []

                # 收集所有学生和用户数据
                students_to_add = []
                users_to_add = []
                student_user_mapping = []  # 用于关联学生和用户

                for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
                    try:
                        # 将行数据转换为字典（表头->值）
                        row_data = dict(zip(headers, row))

                        # 提取必要字段并确保是字符串类型
                        name = str(row_data.get('姓名', '')) if row_data.get('姓名') is not None else ''
                        student_number = str(row_data.get('学号', '')) if row_data.get('学号') is not None else ''
                        username = str(row_data.get('用户名', '')) if row_data.get('用户名') is not None else ''
                        password = str(row_data.get('密码', '')) if row_data.get('密码') is not None else ''

                        # 验证必要字段
                        if not all([name.strip(), student_number.strip(), username.strip(), password.strip()]):
                            fail_count += 1
                            fail_reasons.append(f'行{row_idx}：缺少必要信息')
                            continue

                        # 检查学号和用户名是否已存在
                        if Student.query.filter_by(student_number=student_number).first():
                            fail_count += 1
                            fail_reasons.append(f'行{row_idx}：学号{student_number}已存在')
                            continue

                        if User.query.filter_by(username=username).first():
                            fail_count += 1
                            fail_reasons.append(f'行{row_idx}：用户名{username}已存在')
                            continue

                        # 创建学生记录对象（暂不提交到数据库）
                        student = Student(
                            name=name.strip(),
                            student_number=student_number.strip(),
                            gender=str(row_data.get('性别')) if row_data.get('性别') is not None else None,
                            age=int(row_data.get('年龄')) if row_data.get('年龄') is not None and str(row_data.get('年龄')).isdigit() else None,
                            major=str(row_data.get('专业')) if row_data.get('专业') is not None else None,
                            email=str(row_data.get('邮箱')) if row_data.get('邮箱') is not None else None
                        )

                        students_to_add.append((student, password.strip(), username.strip()))
                        success_count += 1

                    except Exception as row_error:
                        fail_count += 1
                        fail_reasons.append(f'行{row_idx}：处理失败 - {str(row_error)}')
                        continue

                # 批量插入学生记录
                if students_to_add:
                    try:
                        # 批量添加所有学生
                        student_objects = [item[0] for item in students_to_add]
                        db.session.add_all(student_objects)
                        db.session.flush()  # 刷新以获取所有学生的ID

                        # 创建对应的用户记录
                        users_to_add = []
                        for student, password, username in students_to_add:
                            user = User(
                                username=username,
                                role='student',
                                student_id=student.id
                            )
                            user.set_password(password)
                            users_to_add.append(user)

                        # 批量添加所有用户
                        db.session.add_all(users_to_add)
                        db.session.commit()

                        logger.info(f"批量导入完成：{len(student_objects)}个学生，{len(users_to_add)}个用户")

                    except Exception as batch_error:
                        db.session.rollback()
                        # 重置计数器
                        fail_count += success_count
                        success_count = 0
                        fail_reasons.append(f'批量提交失败：{str(batch_error)}')

                # 显示导入结果
                flash(f'批量导入完成！成功：{success_count} 条，失败：{fail_count} 条')
                if fail_reasons:
                    for reason in fail_reasons:
                        flash(f'失败原因：{reason}')

            except Exception as e:
                db.session.rollback()
                logger.error(f"批量导入失败: {str(e)}", exc_info=True)
                flash(f'导入失败：{str(e)}')
            return redirect(url_for('main.manage_students'))
        if action == 'delete' and student_id:
            student = Student.query.get(student_id)
            if student:
                # 删除关联的用户
                user = User.query.filter_by(student_id=student_id).first()
                if user:
                    db.session.delete(user)
                db.session.delete(student)
                db.session.commit()
                flash('Student deleted successfully')

        elif action == 'add':
            name = request.form.get('name')
            student_number = request.form.get('student_number')
            gender = request.form.get('gender')
            age = request.form.get('age')
            major = request.form.get('major')
            email = request.form.get('email')
            username = request.form.get('username')
            password = request.form.get('password')

            if Student.query.filter_by(student_number=student_number).first():
                flash('Student number already exists')
            elif User.query.filter_by(username=username).first():
                flash('Username already exists')
            else:
                student = Student(
                    name=name,
                    student_number=student_number,
                    gender=gender,
                    age=int(age) if age else None,
                    major=major,
                    email=email
                )
                db.session.add(student)
                db.session.commit()

                # 创建学生用户
                user = User(
                    username=username,
                    role='student',
                    student_id=student.id
                )
                user.set_password(password)
                db.session.add(user)
                db.session.commit()

                flash('Student added successfully')

        elif action == 'edit' and student_id:
            student = Student.query.get(student_id)
            if student:
                student.name = request.form.get('name')
                student.student_number = request.form.get('student_number')
                student.gender = request.form.get('gender')
                student.age = int(request.form.get('age')) if request.form.get('age') else None
                student.major = request.form.get('major')
                student.email = request.form.get('email')
                db.session.commit()
                flash('Student updated successfully')

    students = Student.query.all()
    return render_template('admin/manage_students.html', students=students)



# 教师路由
@main.route('/teacher/dashboard', methods=['GET', 'POST'])
@login_required
@check_session_timeout  # 新增：添加超时检查
@role_required('teacher')
def teacher_dashboard():
    # if not current_user.is_teacher():
    #     flash('Access denied')
    #     return redirect(url_for('main.index'))

    students = Student.query.all()
    return render_template('teacher/dashboard.html', students=students)


# 学生路由
@main.route('/student/profile')
@login_required
@check_session_timeout  # 新增：添加超时检查
@role_required('student')
def student_profile():
    # if not current_user.is_student():
    #     flash('Access denied')
    #     return redirect(url_for('main.index'))

    # 修正：直接获取关联的学生对象
    student = current_user.student
    if not student:
        flash('Student information not found')
        return redirect(url_for('main.logout'))

    return render_template('student/profile.html', student=student)
